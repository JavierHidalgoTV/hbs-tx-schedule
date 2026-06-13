"""Extract FWC26_TX_Input_v3.xlsx into JS literals for embedding in the HTML."""
import json, re, sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

XLSX = "FWC26_TX_Input_v3.xlsx"
wb = openpyxl.load_workbook(XLSX, data_only=True)

def sheet_rows(name, skip=2):
    """Return list-of-lists, skipping `skip` header rows, stripping fully-empty rows."""
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < skip:
            continue
        vals = list(row)
        if any(v is not None and str(v).strip() != "" for v in vals):
            rows.append(vals)
    return rows

def cell_str(v):
    if v is None: return None
    return str(v).strip() or None

def cell_num(v):
    if v is None: return None
    try: return float(v)
    except: return None

def excel_date_to_iso(v):
    """Convert Excel serial date or DD/MM/YYYY string → YYYY-MM-DD."""
    if v is None: return None
    if isinstance(v, (int, float)):
        # Excel serial: days since 1900-01-00 (with leap-year bug)
        base = datetime(1899, 12, 30)
        d = base + timedelta(days=int(v))
        return d.strftime("%Y-%m-%d")
    s = str(v).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    # Try YYYY-MM-DD directly
    m2 = re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
    if m2: return s[:10]
    return None

def date_label(iso):
    if not iso: return ''
    y, mo, d = iso.split('-')
    months = ['January','February','March','April','May','June',
              'July','August','September','October','November','December']
    return f"{int(d)} {months[int(mo)-1]} {y}"

def hhmm_to_decimal(v):
    """Convert Excel time serial (0–1) or HH:MM string → decimal hours."""
    if v is None: return None
    if isinstance(v, float) and 0 <= v <= 1:
        return round(v * 24, 4)
    s = str(v).strip()
    m = re.match(r'^(\d{1,2}):(\d{2})', s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60
    return None

def to_ibc_str(v):
    """Convert time value → 'HH:MM IBC' string."""
    if v is None or str(v).strip() == '': return None
    s = str(v).strip()
    if s.endswith('IBC'): return s
    if isinstance(v, float) and 0 <= v <= 1:
        total = round(v * 24 * 60)
        s = f"{total//60:02d}:{total%60:02d}"
    else:
        # strip seconds HH:MM:SS → HH:MM
        s = re.sub(r'^(\d{1,2}:\d{2}):\d{2}$', r'\1', s)
    return s + ' IBC'

def slugify(s):
    return re.sub(r'[^a-z0-9]', '', s.lower())[:10] or 'svc'

# ── Settings ──────────────────────────────────────────────────────────────────
settings = {}
if 'Settings' in wb.sheetnames:
    for r in sheet_rows('Settings'):
        k = cell_str(r[0])
        if k:
            settings[k] = r[1]

match_dur_min = int(cell_num(settings.get('match_duration_min')) or 105)

# ── Services / Feeds / Subtasks ───────────────────────────────────────────────
svc_map = {}  # name → service dict

if 'Services' in wb.sheetnames:
    for i, r in enumerate(sheet_rows('Services')):
        name = cell_str(r[0])
        if not name: continue
        svc_map[name] = {
            'id': slugify(name) + '_' + str(i),
            'name': name,
            'start': int(cell_num(r[1]) or 0),
            'stop':  int(cell_num(r[2]) or 0),
            'feeds': []
        }

if 'Feeds' in wb.sheetnames:
    for r in sheet_rows('Feeds'):
        svc_name = cell_str(r[0])
        feed_name = cell_str(r[1])
        if not svc_name or not feed_name or svc_name not in svc_map: continue
        is_indiv = bool(r[4])
        af = cell_str(r[5])
        feed = {'name': feed_name, 'subtasks': []}
        if is_indiv:
            feed['individual'] = True
        else:
            feed['start'] = int(cell_num(r[2]) or 0)
            feed['stop']  = int(cell_num(r[3]) or 0)
        if af:
            feed['activeFrom'] = af
        svc_map[svc_name]['feeds'].append(feed)

if 'Subtasks' in wb.sheetnames:
    for r in sheet_rows('Subtasks'):
        svc_name  = cell_str(r[0])
        feed_name = cell_str(r[1])
        st_name   = cell_str(r[2])
        offset    = cell_num(r[3])
        if not svc_name or not feed_name or not st_name or offset is None: continue
        if svc_name not in svc_map: continue
        svc = svc_map[svc_name]
        feed = next((f for f in svc['feeds'] if f['name'] == feed_name), None)
        if feed:
            feed['subtasks'].append({'name': st_name, 'offset': int(offset)})

services = [s for s in svc_map.values() if s['feeds']]

# ── Matches ───────────────────────────────────────────────────────────────────
match_sheet = 'Matches' if 'Matches' in wb.sheetnames else wb.sheetnames[0]
raw_rows = list(wb[match_sheet].iter_rows(values_only=True))

# Find first data row: col 0 is a positive integer
data_start = 0
for i, row in enumerate(raw_rows[:6]):
    v = row[0]
    if isinstance(v, (int, float)) and v > 0 and float(v) == int(v):
        data_start = i; break
    if isinstance(v, str) and re.match(r'^\d+$', v.strip()):
        data_start = i; break

matches = []
for row in raw_rows[data_start:]:
    if not row or all(v is None or str(v).strip() == '' for v in row): continue
    iso = excel_date_to_iso(row[2])
    if not iso: continue
    ko_ibc = hhmm_to_decimal(row[9])
    if ko_ibc is None: continue
    matches.append({
        'm':         int(cell_num(row[0]) or (len(matches)+1)),
        'md':        int(cell_num(row[1]) or 0),
        'date':      iso,
        'dateLabel': date_label(iso),
        'ch':        cell_str(row[3]) or 'Match A',
        'teamA':     cell_str(row[4]) or '???',
        'teamB':     cell_str(row[5]) or '???',
        'fullA':     cell_str(row[6]) or '???',
        'fullB':     cell_str(row[7]) or '???',
        'city':      cell_str(row[8]) or '',
        'koIBC':     ko_ibc,
        'dur':       int(cell_num(row[10]) or match_dur_min),
        'trainA':    to_ibc_str(row[11]),
        'pcA':       to_ibc_str(row[12]),
        'trainB':    to_ibc_str(row[13]),
        'pcB':       to_ibc_str(row[14]),
    })

# ── Output JS literals ────────────────────────────────────────────────────────
def to_js(obj, indent=0):
    """Compact but readable JSON → JS literal (no quoted keys)."""
    return json.dumps(obj, ensure_ascii=False, separators=(', ', ': '))

svc_js_lines = []
for s in services:
    feeds_js = json.dumps(s['feeds'], ensure_ascii=False, separators=(', ', ':'))
    line = (f"  {{id:{json.dumps(s['id'])}, name:{json.dumps(s['name'])}, "
            f"start:{s['start']}, stop:{s['stop']}, feeds:{feeds_js}}}")
    svc_js_lines.append(line)

svc_block = "const BAKED_SERVICES=[\n" + ",\n".join(svc_js_lines) + "\n];"

match_js_lines = []
for m in matches:
    def jv(v): return json.dumps(v, ensure_ascii=False) if isinstance(v, str) else (str(v) if v is not None else 'null')
    line = (f"  {{m:{m['m']},md:{m['md']},date:{jv(m['date'])},dateLabel:{jv(m['dateLabel'])},"
            f"ch:{jv(m['ch'])},teamA:{jv(m['teamA'])},teamB:{jv(m['teamB'])},"
            f"fullA:{jv(m['fullA'])},fullB:{jv(m['fullB'])},city:{jv(m['city'])},"
            f"koIBC:{m['koIBC']},dur:{m['dur']},"
            f"trainA:{jv(m['trainA'])},pcA:{jv(m['pcA'])},"
            f"trainB:{jv(m['trainB'])},pcB:{jv(m['pcB'])}}}")
    match_js_lines.append(line)

match_block = "const BAKED_MATCHES=[\n" + ",\n".join(match_js_lines) + "\n];"

with open("_baked_data.js", "w", encoding="utf-8") as f:
    f.write(svc_block + "\n\n" + match_block + "\n")

print(f"Done. {len(services)} services, {len(matches)} matches -> _baked_data.js")
print(f"Settings: {settings}")
