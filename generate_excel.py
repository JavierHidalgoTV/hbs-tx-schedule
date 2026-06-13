# -*- coding: utf-8 -*-
"""Generate FWC26_TX_Input_v3.xlsx — 104 matches, IBC time only (no UTC columns)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
wb.remove(wb.active)

H_FILL = PatternFill("solid", fgColor="1E2235")
H_FONT = Font(bold=True, color="DDE1ED", name="Calibri", size=10)
D_FONT  = Font(name="Calibri", size=10)
AF_FILL = PatternFill("solid", fgColor="1A2E0A")
AF_FONT = Font(color="8BC34A", name="Calibri", size=10, bold=True)

def style_header(ws, row=2):
    for cell in ws[row]:
        if cell.value:
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")

def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

# ─── Services ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Services")
ws.append(["SERVICES — Start offset: min from KO. Stop offset: min from Final Whistle (FW = KO + match_duration_min). Negative stop = before FW; positive = after FW."])
ws.append(["Service Name", "Start Offset\n(min from KO)", "Stop Offset\n(min from FW)", "Notes"])
style_header(ws)
for row in [
    ("Digital Content Live",             -150,  65,  None),
    ("DCL AntiPiracy",                    -90,  15,  None),
    ("FCO",                               -45,  15,  None),
    ("RIS",                               -45,  60,  None),
    ("Datatainment",                      -45,  15,  None),
    ("Match Buddy",                       -150,  65,  None),
    ("Automated Content Creation (WSC)", -150,  65,  None),
    ("Closed Captions",                   -90,  15,  None),
]:
    ws.append(list(row))
set_widths(ws, [36, 14, 14, 30])

# ─── Feeds ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Feeds")
ws.append(["FEEDS — TX Start: minutes from KO (negative). TX End: minutes from Final Whistle (FW = KO + match_duration_min). Negative TX End = before FW; positive = after FW."])
ws.append(["Service", "Feed Name", "TX Start\n(min from KO)", "TX End\n(min from FW)", "Individual\nSchedule", "active_from\n(DD/MM/YYYY)", "Notes"])
style_header(ws)
feeds = [
    ("Digital Content Live", "Team A",        -150, 65,  None, None, None),
    ("Digital Content Live", "Team B",        -150, 65,  None, None, None),
    ("Digital Content Live", "Fan & Reaction",-120, 60,  None, None, None),
    ("Digital Content Live", "WF",             -90, 15,  None, None, None),
    ("Digital Content Live", "ESF",            -90, 15,  None, None, None),
    ("Digital Content Live", "MD-1 PC",        None,None,"YES",None, None),
    ("Digital Content Live", "MD-1 Training",  None,None,"YES",None, None),
    ("DCL AntiPiracy",       "ESF",            -90, 15,  None, None, None),
    ("DCL AntiPiracy",       "WF",             -90, 15,  None, None, None),
    ("FCO",                  "WF",             -90, 15,  None, None, None),
    ("RIS",                  "WF",             -90, 15,  None, None, None),
    ("RIS",                  "Fan & Reaction",-120, 60,  None, None, None),
    ("Datatainment",         "Datatainment",   -90, 15,  None, None, None),
    ("Match Buddy",          "MP UNI",         -60, 15,  None, None, None),
    ("Match Buddy",          "WF",             -90, 15,  None, None, None),
    ("Match Buddy",          "ISO 9 Tactical Main", -60, 20, None, None, None),
    ("Match Buddy",          "ISO 10 Player A",     -10, 10, None, "13/06/2026", "Active from MD3 onwards"),
    ("Match Buddy",          "ISO 11 Player B",     -10, 10, None, "13/06/2026", "Active from MD3 onwards"),
    ("Match Buddy",          "Fan & Reaction",-120, 60,  None, None, None),
    ("Match Buddy",          "Team A",        -150, 65,  None, None, None),
    ("Match Buddy",          "Team B",        -150, 65,  None, None, None),
    ("Match Buddy",          "ISO 1",          -30, 20,  None, None, None),
    ("Match Buddy",          "ISO 6 Cable Cam",-70, 20,  None, None, None),
    ("Automated Content Creation (WSC)", "Team A",        -150, 65,  None, None, None),
    ("Automated Content Creation (WSC)", "Team B",        -150, 65,  None, None, None),
    ("Automated Content Creation (WSC)", "Fan & Reaction",-120, 60,  None, None, None),
    ("Automated Content Creation (WSC)", "WF",             -90, 15,  None, None, None),
    ("Automated Content Creation (WSC)", "MD-1 PC",        None,None,"YES",None, None),
    ("Automated Content Creation (WSC)", "MD-1 Training",  None,None,"YES",None, None),
    ("Closed Captions",      "WF",             -90, 120, None, None, None),
]
for f in feeds:
    ws.append(list(f))
for row in ws.iter_rows(min_row=3):
    if row[5].value:
        row[5].fill = AF_FILL
        row[5].font = AF_FONT
set_widths(ws, [36, 22, 14, 10, 12, 14, 28])

# ─── Subtasks ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Subtasks")
ws.append(["SUBTASKS — One row per subtask per feed. Start Offset = minutes from KO (negative = before KO)."])
ws.append(["Service", "Feed", "Subtask Name", "Start Offset\n(min from KO)", "Notes"])
style_header(ws)
for row in [
    ("FCO",  "WF",             "Frontend Available to FIFA",  -15, None),
    ("FCO",  "WF",             "SRT stream available",        -45, None),
    ("RIS",  "WF",             "WebRTC available to MP's",    -15, None),
    ("RIS",  "WF",             "SRT available to MP's",       -45, None),
    ("RIS",  "Fan & Reaction", "WebRTC available to MP's",    -60, None),
    ("RIS",  "Fan & Reaction", "SRT available to MP's",       -45, None),
]:
    ws.append(list(row))
set_widths(ws, [36, 22, 32, 14, 30])

# ─── Settings ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Settings")
ws.append(["SETTINGS — App configuration. Do not change Key names."])
ws.append(["Key", "Value", "Description"])
style_header(ws)
for row in [
    ("ibc_offset",         -5,            "IBC Time offset from UTC. CDT=-5. Change to -6 for CST."),
    ("tour_start",         "10/06/2026",  "Calendar start date (DD/MM/YYYY)"),
    ("tour_end",           "19/07/2026",  "Calendar end date (DD/MM/YYYY)"),
    ("match_duration_min", 105,           "KO → Final Whistle (min). Feed/service TX End offsets are relative to this. Override per match with 'Dur (min)' column."),
    ("md1_train_dur_min",  90,            "MD-1 Training block duration estimate (minutes)"),
    ("md1_pc_dur_min",     45,            "MD-1 Press Conference block duration (minutes)"),
    ("px_per_hour",        160,           "Pixels per hour in 1h zoom mode"),
]:
    ws.append(list(row))
set_widths(ws, [22, 14, 50])

# ─── Matches ─────────────────────────────────────────────────────────────────
# Columns: M#, MD#, Date, Channel, TriA, TriB, FullA, FullB, City,
#          IBC KO, [Dur inserted], Train A, PC A, Train B, PC B
ws = wb.create_sheet("Matches")
ws.append(["MATCHES — One row per match. All times in IBC (CDT = UTC-5). MD-1 times are absolute IBC HH:MM."])
ws.append([
    "M#", "MD#", "Date\n(DD/MM/YYYY)", "Channel\n(Match A-F)",
    "Team A\n(trigram)", "Team B\n(trigram)", "Full Name A", "Full Name B",
    "City", "IBC KO\n(HH:MM)", "Dur (min)",
    "MD-1\nTrain A (IBC)", "MD-1\nPC A (IBC)",
    "MD-1\nTrain B (IBC)", "MD-1\nPC B (IBC)",
])
style_header(ws)
# Tuple: (M#, MD#, Date, Channel, TriA, TriB, FullA, FullB, City, IBC_KO, TrainA, PCA, TrainB, PCB)
# Dur (min) is always None (use global default from Settings).
# MD-1 times = None for knockout rounds (no scheduled training/PC).
matches = [
    # MD1 — 11 Jun
    (  1, 1,'11/06/2026','Match A','MEX','RSA','Mexico','South Africa','Mexico City','14:00','12:00','15:30','18:00','16:30'),
    (  2, 1,'11/06/2026','Match B','KOR','CZE','Korea Republic','Czechia','Guadalajara','21:00','17:30','15:30','18:45','16:30'),
    # MD2 — 12 Jun
    (  3, 2,'12/06/2026','Match A','CAN','BIH','Canada','Bosnia and Herzegovina','Toronto','14:00','14:00','17:00','10:00','17:45'),
    (  4, 2,'12/06/2026','Match B','USA','PAR','United States','Paraguay','Los Angeles, CA','20:00','13:00','17:00','11:00','20:45'),
    # MD3 — 13 Jun
    (  5, 3,'13/06/2026','Match C','HAI','SCO','Haiti','Scotland','Boston, MA','20:00','17:00','16:00','09:00','11:15'),
    (  6, 3,'13/06/2026','Match D','AUS','TUR','Australia','Türkiye','Vancouver','23:00','21:00','19:30','17:00','20:30'),
    (  7, 3,'13/06/2026','Match B','BRA','MAR','Brazil','Morocco','New York, NJ','17:00','10:00','15:15','17:00','14:30'),
    (  8, 3,'13/06/2026','Match A','QAT','SUI','Qatar','Switzerland','San Francisco, CA','14:00','12:00','17:45','12:15','20:45'),
    # MD4 — 14 Jun
    (  9, 4,'14/06/2026','Match C','CIV','ECU',"Côte d'Ivoire",'Ecuador','Philadelphia, PA','18:00','16:00','14:45','08:00','17:45'),
    ( 10, 4,'14/06/2026','Match A','GER','CUW','Germany','Curaçao','Houston, TX','12:00','10:00','18:45','09:00','19:45'),
    ( 11, 4,'14/06/2026','Match B','NED','JPN','Netherlands','Japan','Dallas, TX','15:00','10:30','18:45','10:00','15:45'),
    ( 12, 4,'14/06/2026','Match D','SWE','TUN','Sweden','Tunisia','Monterrey','21:00','19:00','15:30','21:00','16:45'),
    # MD5 — 15 Jun
    ( 13, 5,'15/06/2026','Match C','KSA','URU','Saudi Arabia','Uruguay','Miami, FL','17:00','17:00','14:45','09:00','17:45'),
    ( 14, 5,'15/06/2026','Match A','ESP','CPV','Spain','Cabo Verde','Atlanta, GA','11:00','09:30','14:45','13:00','17:45'),
    ( 15, 5,'15/06/2026','Match D','IRN','NZL','IR Iran','New Zealand','Los Angeles, CA','20:00','19:30','17:45','12:45','20:45'),
    ( 16, 5,'15/06/2026','Match B','BEL','EGY','Belgium','Egypt','Seattle, WA','14:00','13:00','16:30','12:00','17:30'),
    # MD6 — 16 Jun
    ( 17, 6,'16/06/2026','Match A','FRA','SEN','France','Senegal','New York, NJ','14:00','15:00','09:30','16:00','13:30'),
    ( 18, 6,'16/06/2026','Match B','IRQ','NOR','Iraq','Norway','Boston, MA','17:00','17:30','19:00','14:30','16:00'),
    ( 19, 6,'16/06/2026','Match C','ARG','ALG','Argentina','Algeria','Kansas City, MO','20:00','18:00','14:30','20:00','15:30'),
    ( 20, 6,'16/06/2026','Match D','AUT','JOR','Austria','Jordan','San Francisco, CA','23:00','13:00','20:45','22:00','17:45'),
    # MD7 — 17 Jun
    ( 21, 7,'17/06/2026','Match C','GHA','PAN','Ghana','Panama','Toronto','18:00','18:00','13:30','08:00','14:30'),
    ( 22, 7,'17/06/2026','Match B','ENG','CRO','England','Croatia','Dallas, TX','15:00','11:00','18:45','15:00','19:45'),
    ( 23, 7,'17/06/2026','Match A','POR','COD','Portugal','Congo DR','Houston, TX','12:00','09:30','18:45','13:00','15:45'),
    ( 24, 7,'17/06/2026','Match D','UZB','COL','Uzbekistan','Colombia','Mexico City','21:00','09:30','19:45','18:30','16:45'),
    # MD8 — 18 Jun
    ( 25, 8,'18/06/2026','Match A','CZE','RSA','Czechia','South Africa','Atlanta, GA','11:00','09:00','13:30','16:00','14:30'),
    ( 26, 8,'18/06/2026','Match B','SUI','BIH','Switzerland','Bosnia and Herzegovina','Los Angeles, CA','14:00','12:45','20:45','11:00','21:45'),
    ( 27, 8,'18/06/2026','Match C','CAN','QAT','Canada','Qatar','Vancouver','17:00','17:00','20:45','13:00','17:30'),
    ( 28, 8,'18/06/2026','Match D','MEX','KOR','Mexico','Korea Republic','Guadalajara','20:00','19:00','15:30','17:30','16:30'),
    # MD9 — 19 Jun
    ( 29, 9,'19/06/2026','Match C','BRA','HAI','Brazil','Haiti','Philadelphia, PA','19:30','09:00','17:45','16:30','18:45'),
    ( 30, 9,'19/06/2026','Match B','SCO','MAR','Scotland','Morocco','Boston, MA','17:00','09:00','10:30','17:00','16:15'),
    ( 31, 9,'19/06/2026','Match D','TUR','PAR','Türkiye','Paraguay','San Francisco, CA','22:00','19:45','16:30','20:30','17:30'),
    ( 32, 9,'19/06/2026','Match A','USA','AUS','United States','Australia','Seattle, WA','14:00','13:00','17:45','14:00','20:45'),
    # MD10 — 20 Jun
    ( 33,10,'20/06/2026','Match B','GER','CIV','Germany',"Côte d'Ivoire",'Toronto','15:00','10:00','17:45','16:00','14:45'),
    ( 34,10,'20/06/2026','Match C','ECU','CUW','Ecuador','Curaçao','Kansas City, MO','19:00','08:00','18:45','09:00','19:45'),
    ( 35,10,'20/06/2026','Match A','NED','SWE','Netherlands','Sweden','Houston, TX','12:00','10:30','18:45','09:30','15:45'),
    ( 36,10,'20/06/2026','Match D','TUN','JPN','Tunisia','Japan','Monterrey','23:00','23:00','15:30','19:00','16:30'),
    # MD11 — 21 Jun
    ( 37,11,'21/06/2026','Match C','URU','CPV','Uruguay','Cabo Verde','Miami, FL','17:00','09:00','17:45','08:00','18:45'),
    ( 38,11,'21/06/2026','Match A','ESP','KSA','Spain','Saudi Arabia','Atlanta, GA','11:00','10:00','13:30','17:00','14:30'),
    ( 39,11,'21/06/2026','Match B','BEL','IRN','Belgium','IR Iran','Los Angeles, CA','14:00','13:00','16:30','18:30','20:45'),
    ( 40,11,'21/06/2026','Match D','NZL','EGY','New Zealand','Egypt','Vancouver','20:00','12:45','16:30','20:00','17:30'),
    # MD12 — 22 Jun
    ( 41,12,'22/06/2026','Match C','NOR','SEN','Norway','Senegal','New York, NJ','19:00','09:00','17:45','15:00','13:30'),
    ( 42,12,'22/06/2026','Match B','FRA','IRQ','France','Iraq','Philadelphia, PA','16:00','16:00','14:30','19:30','13:30'),
    ( 43,12,'22/06/2026','Match A','ARG','AUT','Argentina','Austria','Dallas, TX','12:00','10:00','18:45','11:00','15:45'),
    ( 44,12,'22/06/2026','Match D','JOR','ALG','Jordan','Algeria','San Francisco, CA','22:00','22:00','16:30','19:00','17:30'),
    # MD13 — 23 Jun
    ( 45,13,'23/06/2026','Match B','ENG','GHA','England','Ghana','Boston, MA','15:00','10:15','17:45','15:00','18:45'),
    ( 46,13,'23/06/2026','Match C','PAN','CRO','Panama','Croatia','Toronto','18:00','07:30','13:30','18:00','14:15'),
    ( 47,13,'23/06/2026','Match A','POR','UZB','Portugal','Uzbekistan','Houston, TX','12:00','09:30','18:45','09:30','19:45'),
    ( 48,13,'23/06/2026','Match D','COL','COD','Colombia','Congo DR','Guadalajara','21:00','18:00','16:45','10:00','19:45'),
    # MD14 — 24 Jun
    ( 49,14,'24/06/2026','Match C','SCO','BRA','Scotland','Brazil','Miami, FL','17:00','09:00','13:30','09:00','17:45'),
    ( 50,14,'24/06/2026','Match D','MAR','HAI','Morocco','Haiti','Atlanta, GA','17:00','17:00','13:30','10:00','14:45'),
    ( 51,14,'24/06/2026','Match A','SUI','CAN','Switzerland','Canada','Vancouver','14:00','13:00','16:30','12:00','17:30'),
    ( 52,14,'24/06/2026','Match B','BIH','QAT','Bosnia and Herzegovina','Qatar','Seattle, WA','14:00','11:00','20:45','13:00','17:30'),
    ( 53,14,'24/06/2026','Match E','CZE','MEX','Czechia','Mexico','Mexico City','20:00','18:00','15:30','19:30','16:30'),
    ( 54,14,'24/06/2026','Match F','RSA','KOR','South Africa','Korea Republic','Monterrey','20:00','18:00','15:30','21:30','16:45'),
    # MD15 — 25 Jun
    ( 55,15,'25/06/2026','Match A','CUW','CIV','Curaçao',"Côte d'Ivoire",'Philadelphia, PA','15:00','09:00','17:45','14:00','18:45'),
    ( 56,15,'25/06/2026','Match B','ECU','GER','Ecuador','Germany','New York, NJ','15:00','08:00','17:45','10:00','18:45'),
    ( 57,15,'25/06/2026','Match C','JPN','SWE','Japan','Sweden','Dallas, TX','18:00','10:00','14:30','18:00','15:30'),
    ( 58,15,'25/06/2026','Match D','TUN','NED','Tunisia','Netherlands','Kansas City, MO','18:00','18:00','14:30','10:30','15:30'),
    ( 59,15,'25/06/2026','Match E','TUR','USA','Türkiye','United States','Los Angeles, CA','21:00','19:30','16:30','13:00','17:45'),
    ( 60,15,'25/06/2026','Match F','PAR','AUS','Paraguay','Australia','San Francisco, CA','21:00','19:30','16:30','21:00','17:30'),
    # MD16 — 26 Jun
    ( 61,16,'26/06/2026','Match A','NOR','FRA','Norway','France','Boston, MA','14:00','16:30','18:00','14:30','09:00'),
    ( 62,16,'26/06/2026','Match B','SEN','IRQ','Senegal','Iraq','Toronto','14:00','16:00','13:30','11:00','14:30'),
    ( 63,16,'26/06/2026','Match E','EGY','IRN','Egypt','IR Iran','Seattle, WA','22:00','22:00','16:30','18:30','20:45'),
    ( 64,16,'26/06/2026','Match F','NZL','BEL','New Zealand','Belgium','Vancouver','22:00','12:45','17:45','13:00','20:45'),
    ( 65,16,'26/06/2026','Match C','CPV','KSA','Cabo Verde','Saudi Arabia','Houston, TX','19:00','08:00','18:45','17:30','15:45'),
    ( 66,16,'26/06/2026','Match D','URU','ESP','Uruguay','Spain','Guadalajara','19:00','09:00','19:45','12:00','16:45'),
    # MD17 — 27 Jun
    ( 67,17,'27/06/2026','Match A','PAN','ENG','Panama','England','New York, NJ','16:00','09:00','14:45','10:15','17:45'),
    ( 68,17,'27/06/2026','Match B','CRO','GHA','Croatia','Ghana','Philadelphia, PA','16:00','16:00','18:00','20:00','14:30'),
    ( 69,17,'27/06/2026','Match E','ALG','AUT','Algeria','Austria','Kansas City, MO','21:00','18:00','15:30','19:00','14:30'),
    ( 70,17,'27/06/2026','Match F','JOR','ARG','Jordan','Argentina','Dallas, TX','21:00','20:15','15:45','16:45','18:45'),
    ( 71,17,'27/06/2026','Match C','COL','POR','Colombia','Portugal','Miami, FL','18:30','16:00','13:30','17:30','14:30'),
    ( 72,17,'27/06/2026','Match D','COD','UZB','Congo DR','Uzbekistan','Atlanta, GA','18:30','18:00','13:30','10:00','14:30'),
    # MD18-34 — Knockout rounds (no MD-1 training/PC times)
    ( 73,18,'28/06/2026','Match A','2A','2B','2A','2B','Los Angeles, CA','14:00',None,None,None,None),
    ( 74,19,'29/06/2026','Match B','1E','3ABCDF','1E','3ABCDF','Boston, MA','15:30',None,None,None,None),
    ( 75,19,'29/06/2026','Match C','1F','2C','1F','2C','Monterrey','20:00',None,None,None,None),
    ( 76,19,'29/06/2026','Match A','1C','2F','1C','2F','Houston, TX','12:00',None,None,None,None),
    ( 77,20,'30/06/2026','Match B','1I','3CDFGH','1I','3CDFGH','New York, NJ','16:00',None,None,None,None),
    ( 78,20,'30/06/2026','Match A','2E','2I','2E','2I','Dallas, TX','12:00',None,None,None,None),
    ( 79,20,'30/06/2026','Match C','1A','3CEFHI','1A','3CEFHI','Mexico City','20:00',None,None,None,None),
    ( 80,21,'01/07/2026','Match A','1L','3EHIJK','1L','3EHIJK','Atlanta, GA','11:00',None,None,None,None),
    ( 81,21,'01/07/2026','Match C','1D','3BEFIJ','1D','3BEFIJ','San Francisco, CA','19:00',None,None,None,None),
    ( 82,21,'01/07/2026','Match B','1G','3AEHIJ','1G','3AEHIJ','Seattle, WA','15:00',None,None,None,None),
    ( 83,22,'02/07/2026','Match B','2K','2L','2K','2L','Toronto','18:00',None,None,None,None),
    ( 84,22,'02/07/2026','Match A','1H','2J','1H','2J','Los Angeles, CA','14:00',None,None,None,None),
    ( 85,22,'02/07/2026','Match C','1B','3EFGHIJ','1B','3EFGHIJ','Vancouver','22:00',None,None,None,None),
    ( 86,23,'03/07/2026','Match B','1J','2H','1J','2H','Miami, FL','17:00',None,None,None,None),
    ( 87,23,'03/07/2026','Match C','1K','3DEIJL','1K','3DEIJL','Kansas City, MO','20:30',None,None,None,None),
    ( 88,23,'03/07/2026','Match A','2D','2G','2D','2G','Dallas, TX','13:00',None,None,None,None),
    ( 89,24,'04/07/2026','Match B','W74','W77','W74','W77','Philadelphia, PA','16:00',None,None,None,None),
    ( 90,24,'04/07/2026','Match A','W73','W75','W73','W75','Houston, TX','12:00',None,None,None,None),
    ( 91,25,'05/07/2026','Match A','W76','W78','W76','W78','New York, NJ','15:00',None,None,None,None),
    ( 92,25,'05/07/2026','Match B','W79','W80','W79','W80','Mexico City','19:00',None,None,None,None),
    ( 93,26,'06/07/2026','Match A','W83','W84','W83','W84','Dallas, TX','14:00',None,None,None,None),
    ( 94,26,'06/07/2026','Match B','W81','W82','W81','W82','Seattle, WA','19:00',None,None,None,None),
    ( 95,27,'07/07/2026','Match A','W86','W88','W86','W88','Atlanta, GA','11:00',None,None,None,None),
    ( 96,27,'07/07/2026','Match B','W85','W87','W85','W87','Vancouver','15:00',None,None,None,None),
    ( 97,28,'09/07/2026','Match A','W89','W90','W89','W90','Boston, MA','15:00',None,None,None,None),
    ( 98,29,'10/07/2026','Match A','W93','W94','W93','W94','Los Angeles, CA','14:00',None,None,None,None),
    ( 99,30,'11/07/2026','Match A','W91','W92','W91','W92','Miami, FL','16:00',None,None,None,None),
    (100,30,'11/07/2026','Match B','W95','W96','W95','W96','Kansas City, MO','20:00',None,None,None,None),
    (101,31,'14/07/2026','Match A','W97','W98','W97','W98','Dallas, TX','14:00',None,None,None,None),
    (102,32,'15/07/2026','Match A','W99','W100','W99','W100','Atlanta, GA','14:00',None,None,None,None),
    (103,33,'18/07/2026','Match A','L101','L102','L101','L102','Miami, FL','16:00',None,None,None,None),
    (104,34,'19/07/2026','Match A','W101','W102','W101','W102','New York, NJ','14:00',None,None,None,None),
]
for m in matches:
    row = list(m)
    row.insert(10, None)  # Dur (min) — blank = use global default from Settings
    ws.append(row)
set_widths(ws, [5,5,14,12,8,8,22,24,20,10,8,12,12,12,12])

# ─── Overrides ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Overrides")
ws.append(["OVERRIDES — Absolute IBC times for specific match+service/feed. Overrides offset defaults. Leave blank to use offsets."])
ws.append(["M#", "Channel", "Service", "Feed", "Override TX Start\n(IBC HH:MM)", "Override TX End\n(IBC HH:MM)", "Reason / Notes"])
style_header(ws)
set_widths(ws, [5, 12, 36, 22, 14, 14, 40])

out = "c:/Users/HBMC Op 01/.vscode/hbs-tx-schedule/hbs-tx-schedule/FWC26_TX_Input_v3.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheets: {wb.sheetnames}")
print(f"Matches: {len(matches)}")
