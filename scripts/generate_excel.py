# -*- coding: utf-8 -*-
"""Generate FWC26_TX_Input_v3.xlsx — Config only.

Sheets: Settings, Services, Feeds, Subtasks, Overrides
Match schedule is loaded separately via the UI paste / upload module.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
wb.remove(wb.active)

H_FILL = PatternFill("solid", fgColor="1E2235")
H_FONT = Font(bold=True, color="DDE1ED", name="Calibri", size=10)
D_FONT  = Font(name="Calibri", size=10)
AF_FILL = PatternFill("solid", fgColor="1A2E0A")
AF_FONT = Font(color="8BC34A", name="Calibri", size=10, bold=True)
SVC_MS_FILL = PatternFill("solid", fgColor="1A1E35")
SVC_MS_FONT = Font(color="7B9FE0", name="Calibri", size=10, italic=True)

def style_header(ws, row=2):
    for cell in ws[row]:
        if cell.value:
            cell.font = H_FONT
            cell.fill = H_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center")

def set_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w


# ─── Settings ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Settings")
ws.append(["SETTINGS — App configuration. Do not change Key names."])
ws.append(["Key", "Value", "Description"])
style_header(ws)
for row in [
    ("ibc_offset",         -5,           "IBC Time offset from UTC. CDT=-5. Change to -6 for CST."),
    ("tour_start",         "10/06/2026", "Calendar start date (DD/MM/YYYY)"),
    ("tour_end",           "19/07/2026", "Calendar end date (DD/MM/YYYY)"),
    ("match_duration_min", 105,          "KO → Final Whistle (min). Feed/service TX End offsets are relative to this."),
    ("md1_train_dur_min",  90,           "MD-1 Training block duration estimate (minutes)"),
    ("md1_pc_dur_min",     45,           "MD-1 Press Conference block duration (minutes)"),
    ("px_per_hour",        160,          "Pixels per hour in 1h zoom mode"),
]:
    ws.append(list(row))
set_widths(ws, [22, 14, 55])


# ─── Services ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Services")
ws.append([
    "SERVICES — Start: min from KO (when operator begins pre-run tasks, left edge of bar). "
    "Available: min from KO (when service is available to stakeholders, colour change on bar). "
    "Stop: min from FW (FW = KO + match_duration_min)."
])
ws.append([
    "Service Name",
    "Start Offset\n(min from KO)",
    "Stop Offset\n(min from FW)",
    "Available Time Offset\n(min from KO)",
    "Notes",
])
style_header(ws)
#                                           name                              start  stop  available  notes
for row in [
    ("Digital Content Live",                -150,  65,   -90,  None),
    ("DCL AntiPiracy",                       -90,  15,   -90,  None),
    ("FCO",                                  -45,  15,   -15,  "Available when FIFA frontend is up"),
    ("RIS",                                  -45,  60,   -15,  "Available when WebRTC up for MPs"),
    ("Datatainment",                         -45,  15,   -45,  None),
    ("Match Buddy",                         -150,  65,   -60,  None),
    ("Automated Content Creation (WSC)",    -150,  65,   -90,  None),
    ("Closed Captions",                      -90,  15,   -90,  None),
]:
    ws.append(list(row))
set_widths(ws, [36, 14, 14, 18, 34])


# ─── Feeds ───────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Feeds")
ws.append([
    "FEEDS — TX Start: minutes from KO (negative). TX End: minutes from Final Whistle. "
    "Individual Schedule feeds use absolute IBC times from the match schedule (not KO offsets)."
])
ws.append([
    "Service", "Feed Name",
    "TX Start\n(min from KO)", "TX End\n(min from FW)",
    "Individual\nSchedule", "active_from\n(DD/MM/YYYY)", "Notes",
])
style_header(ws)
feeds = [
    ("Digital Content Live", "Team A",                  -150,  65,  None, None, None),
    ("Digital Content Live", "Team B",                  -150,  65,  None, None, None),
    ("Digital Content Live", "Fan & Reaction",          -120,  60,  None, None, None),
    ("Digital Content Live", "WF",                       -90,  15,  None, None, None),
    ("Digital Content Live", "ESF",                      -90,  15,  None, None, None),
    ("Digital Content Live", "MD-1 PC",                 None, None, "YES", None, None),
    ("Digital Content Live", "MD-1 Training",           None, None, "YES", None, None),
    ("DCL AntiPiracy",       "ESF",                      -90,  15,  None, None, None),
    ("DCL AntiPiracy",       "WF",                       -90,  15,  None, None, None),
    ("FCO",                  "WF",                       -90,  15,  None, None, None),
    ("RIS",                  "WF",                       -90,  15,  None, None, None),
    ("RIS",                  "Fan & Reaction",          -120,  60,  None, None, None),
    ("Datatainment",         "Datatainment",             -90,  15,  None, None, None),
    ("Match Buddy",          "MP UNI",                   -60,  15,  None, None, None),
    ("Match Buddy",          "WF",                       -90,  15,  None, None, None),
    ("Match Buddy",          "ISO 9 Tactical Main",      -60,  20,  None, None, None),
    ("Match Buddy",          "ISO 10 Player A",          -10,  10,  None, "13/06/2026", "Active from MD3 onwards"),
    ("Match Buddy",          "ISO 11 Player B",          -10,  10,  None, "13/06/2026", "Active from MD3 onwards"),
    ("Match Buddy",          "Fan & Reaction",          -120,  60,  None, None, None),
    ("Match Buddy",          "Team A",                  -150,  65,  None, None, None),
    ("Match Buddy",          "Team B",                  -150,  65,  None, None, None),
    ("Match Buddy",          "ISO 1",                    -30,  20,  None, None, None),
    ("Match Buddy",          "ISO 6 Cable Cam",          -70,  20,  None, None, None),
    ("Automated Content Creation (WSC)", "Team A",      -150,  65,  None, None, None),
    ("Automated Content Creation (WSC)", "Team B",      -150,  65,  None, None, None),
    ("Automated Content Creation (WSC)", "Fan & Reaction",-120,60,  None, None, None),
    ("Automated Content Creation (WSC)", "WF",           -90,  15,  None, None, None),
    ("Automated Content Creation (WSC)", "MD-1 PC",     None, None, "YES", None, None),
    ("Automated Content Creation (WSC)", "MD-1 Training",None,None, "YES", None, None),
    ("Closed Captions",      "WF",                       -90, 120,  None, None, None),
]
for f in feeds:
    ws.append(list(f))
for row in ws.iter_rows(min_row=3):
    if row[5].value:
        row[5].fill = AF_FILL
        row[5].font = AF_FONT
set_widths(ws, [36, 22, 14, 10, 12, 14, 28])


# ─── Subtasks ────────────────────────────────────────────────────────────────
ws = wb.create_sheet("Subtasks")
ws.append([
    "SUBTASKS / MILESTONES — Feed filled: marker on that feed bar. "
    "Feed empty: service-level milestone on the service bar (ALL and SERVICE views only). "
    "Start Offset = minutes from KO (negative = before KO)."
])
ws.append([
    "Service",
    "Feed\n(empty = service milestone)",
    "Name",
    "Start Offset\n(min from KO)",
    "Notes",
])
style_header(ws)

subtasks = [
    # Feed-level subtasks (Feed column filled)
    ("FCO",  "WF",             "Frontend Available to FIFA",  -15, None),
    ("FCO",  "WF",             "SRT stream available",        -45, None),
    ("RIS",  "WF",             "WebRTC available to MP's",    -15, None),
    ("RIS",  "WF",             "SRT available to MP's",       -45, None),
    ("RIS",  "Fan & Reaction", "WebRTC available to MP's",    -60, None),
    ("RIS",  "Fan & Reaction", "SRT available to MP's",       -45, None),
    # Service-level milestones (Feed column intentionally empty)
    ("FCO",  "",               "FCO pre-flight complete",     -45, "Service milestone"),
    ("RIS",  "",               "RIS pre-flight complete",     -45, "Service milestone"),
]

for i, row in enumerate(subtasks):
    ws.append(list(row))
    # highlight service-level milestone rows
    if not row[1]:
        excel_row = i + 3  # row 1=header note, row 2=col headers, row 3+=data
        for cell in ws[excel_row]:
            cell.fill = SVC_MS_FILL
            cell.font  = SVC_MS_FONT

set_widths(ws, [36, 22, 32, 14, 30])


# ─── Overrides ───────────────────────────────────────────────────────────────
ws = wb.create_sheet("Overrides")
ws.append([
    "OVERRIDES — Absolute IBC times for a specific match+service/feed. "
    "Overrides the offset defaults. Leave blank to use offsets."
])
ws.append([
    "M#", "Channel", "Service", "Feed",
    "Override TX Start\n(IBC HH:MM)",
    "Override TX End\n(IBC HH:MM)",
    "Reason / Notes",
])
style_header(ws)
set_widths(ws, [5, 12, 36, 22, 14, 14, 40])


# ─── Save ────────────────────────────────────────────────────────────────────
out = "c:/Users/HBMC Op 01/.vscode/hbs-tx-schedule/hbs-tx-schedule/data/FWC26_TX_Input_v3.xlsx"
wb.save(out)
print(f"Saved : {out}")
print(f"Sheets: {wb.sheetnames}")
